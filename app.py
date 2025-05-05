import base64
from PIL import Image
from io import BytesIO
from flask import Flask, render_template, request, jsonify,redirect, send_file, url_for,flash,session
from classes.equipamentos import EquipamentoCRUD
from classes.funcionario import FuncionarioCRUD
import datetime
import warnings
import pandas as pd
import psycopg2  # pip install psycopg2
from openpyxl import load_workbook
from openpyxl.drawing.image import Image as imge
import psycopg2.extras 
from functools import wraps
from datetime import datetime,timedelta
import uuid
import copy
import os

app = Flask(__name__)
app.secret_key = "appEpi"

DB_HOST = "database-2.cdcogkfzajf0.us-east-1.rds.amazonaws.com"
DB_NAME = "postgres"
DB_USER = "postgres"
DB_PASS = "15512332"

crud = EquipamentoCRUD(DB_NAME, DB_USER, DB_PASS, DB_HOST)
crudFuncionario = FuncionarioCRUD(DB_NAME, DB_USER, DB_PASS, DB_HOST)

warnings.filterwarnings("ignore")

def login_required(func): # Lógica do parâmetro de login_required, onde escolhe quais páginas onde apenas o usuário logado pode acessar
    @wraps(func)
    def wrapper(*args, **kwargs):
        if 'loggedin' not in session:
            return redirect(url_for('login'))
        return func(*args, **kwargs)
    return wrapper

@app.route('/login', methods=['POST','GET'])
def login(): # Lógica de login
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)

    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']

        cur.execute("""SELECT * FROM sistema_epi.tb_usuario WHERE username = %s AND password = %s""", (username, password))
        user = cur.fetchone()

        print(user)
        print(user is not None)

        if user is not None:
            session['loggedin'] = True
            session['user_id'] = user['username']
            print(user['username'])
            return redirect(url_for('rota_solicitacao_material',username=username))
        else:
            flash('Usuário ou Senha inválida', category='error')

    return render_template('login.html')

@app.route('/logout')
def logout(): # Botão de logout
	session.pop('loggedin', None)
	session.pop('id', None)
	session.pop('username', None)
	return redirect(url_for('login'))

@app.route('/', methods=['GET'])
@login_required
def pagina_inicial():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # PRIMEIRA TABELA - Todos os itens solicitados
    s = """SELECT
            solic.id,
            solic.id_solicitacao,
            solic.matricula_solicitante,
            solic.codigo_item,
            solic.quantidade,
            solic.motivo,   
            solic.setor_solicitante,   
            solic.funcionario_recebe,   
            solic.data_solicitada,   
            solic.status_devolucao,   
            hist.status,
            ass.assinatura,
            ass.data_assinatura
        FROM sistema_epi.tb_solicitacoes AS solic
        LEFT JOIN (
            SELECT
                id_solicitacao,
                status,
                ROW_NUMBER() OVER (PARTITION BY id_solicitacao ORDER BY data_modificacao DESC) AS row_num
            FROM sistema_epi.tb_historico_solicitacoes
        ) AS hist ON hist.id_solicitacao = solic.id_solicitacao AND hist.row_num = 1
        LEFT JOIN sistema_epi.tb_assinatura AS ass ON ass.id_solicitacao = solic.id_solicitacao
        WHERE solic.status_devolucao IS NULL;
        """

    tb_solicitacoes = pd.read_sql_query(s, conn)

    tb_solicitacoes['data_solicitada'] = pd.to_datetime(tb_solicitacoes['data_solicitada'])
    tb_solicitacoes['data_solicitada'] = tb_solicitacoes['data_solicitada'].dt.strftime('%d/%m/%Y')

    sql = f"select concat(matricula, ' - ', nome) from requisicao.funcionarios"
    
    tb_sql = pd.read_sql_query(sql, conn)

    mapeamento = dict(zip(tb_sql['concat'].str.extract('(\d+)', expand=False), tb_sql['concat']))

    # Substituir na coluna matricula_solicitante
    tb_solicitacoes['matricula_solicitante'] = tb_solicitacoes['matricula_solicitante'].astype(str)
    tb_solicitacoes['matricula_solicitante'] = tb_solicitacoes['matricula_solicitante'].replace(mapeamento)

    tb_solicitacoes['funcionario_recebe'] = tb_solicitacoes['funcionario_recebe'].astype(str)
    tb_solicitacoes['funcionario_recebe'] = tb_solicitacoes['funcionario_recebe'].replace(mapeamento)

    # Agrupar pelo id_solicitacao e manter apenas a primeira linha de cada grupo
    tb_solicitacoes = tb_solicitacoes.groupby('id_solicitacao').first().reset_index()

    tb_solicitacoes = tb_solicitacoes.sort_values(by='id', ascending=False)

    tb_solicitacoes_list = tb_solicitacoes.values.tolist()

    

    # SEGUNDA TABELA - Identificar troca de equipamentos
    query_troca = """SELECT 
                TO_CHAR(sistema_epi.tb_assinatura.data_assinatura, 'DD/MM/YYYY') AS data_assinatura_formatada,
                sistema_epi.tb_solicitacoes.codigo_item,
                CONCAT(requisicao.funcionarios.matricula, ' - ', requisicao.funcionarios.nome) AS matricula_nome_funcionario_recebe,
                sistema_epi.tb_itens.vida_util - EXTRACT(DAY FROM AGE(DATE_TRUNC('day', NOW()), DATE_TRUNC('day', sistema_epi.tb_assinatura.data_assinatura))) AS diferenca_vida_tempo
            FROM 
                sistema_epi.tb_solicitacoes
            LEFT JOIN 
                sistema_epi.tb_assinatura ON sistema_epi.tb_solicitacoes.id_solicitacao = sistema_epi.tb_assinatura.id_solicitacao
            LEFT JOIN 
                sistema_epi.tb_itens ON sistema_epi.tb_solicitacoes.codigo_item LIKE CONCAT('%', sistema_epi.tb_itens.codigo, '%')
            LEFT JOIN
                requisicao.funcionarios ON sistema_epi.tb_solicitacoes.funcionario_recebe = requisicao.funcionarios.matricula
            WHERE 
                sistema_epi.tb_assinatura.data_assinatura IS NOT NULL AND sistema_epi.tb_solicitacoes.status_devolucao IS NULL;
            """
    
    cur.execute(query_troca)
    lista_itens_assinados = cur.fetchall()

    quantidade_assinados,quantidade_devolucao,quantidade_trabalhadores,quantidade_solicitacoes = cardsPaginaInicial(cur)

    return render_template('home.html',tb_solicitacoes_list=tb_solicitacoes_list,lista_itens_assinados=lista_itens_assinados,quantidade_solicitacoes=quantidade_solicitacoes,
                           quantidade_assinados=quantidade_assinados,quantidade_devolucao=quantidade_devolucao,quantidade_trabalhadores=quantidade_trabalhadores)

@app.route('/receber-assinatura', methods=['POST'])
@login_required
def receber_assinatura():
    try:
        conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        # Obtém os dados do JSON enviado no corpo da solicitação
        data = request.get_json()

        # Obtém o dataURL da assinatura
        id_solicitacao = data.get('id_solicitacao')
        dataURL = data.get('dataURL')
        codigo_item = data.get('codigo_item')
        codigo_nome_funcionario = data.get('nome_funcionario')
        matricula,nome = codigo_nome_funcionario.split(" - ")

        data_devolucao = datetime.now().date()

        cur.execute("INSERT INTO sistema_epi.tb_assinatura (id_solicitacao, assinatura) VALUES (%s, %s)", (id_solicitacao, dataURL))

        cur.execute("INSERT INTO sistema_epi.tb_historico_solicitacoes (id_solicitacao, status, motivo) VALUES (%s, %s, %s)", (id_solicitacao, 'Assinado', 'Assinado'))

        if verifica_existencia_data_devolucao(cur, matricula, codigo_item) == True:

            query = """UPDATE sistema_epi.tb_solicitacoes
                SET data_devolucao = %s
                WHERE id = (
                    SELECT id
                        FROM sistema_epi.tb_solicitacoes
                    WHERE funcionario_recebe = %s AND codigo_item = %s AND data_devolucao IS NULL 
                    ORDER BY id ASC
                    LIMIT 1);"""
        
            cur.execute(query, (data_devolucao,matricula,codigo_item))

        conn.commit()

        cur.close()
        conn.close()

        # Exemplo de resposta de volta para o cliente
        return jsonify({'status': 'success', 'message': 'Assinatura recebida com sucesso!'})

    except Exception as e:
        return jsonify({'status': 'error', 'message': str(e)})

@app.route('/dados-execucao', methods=['POST'])
@login_required
def dados_execucao():
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    
    cur = conn.cursor()

    data = request.get_json()

    id_solicitante = data.get('id_solicitante')

    query = f"""
                SELECT
                    solic.*,
                    ass.data_assinatura
                FROM sistema_epi.tb_solicitacoes AS solic
                LEFT JOIN sistema_epi.tb_assinatura AS ass ON ass.id_solicitacao = solic.id_solicitacao
                WHERE solic.id_solicitacao = '{id_solicitante}'
                ORDER BY id asc;
            """
    
    query_solicitacoes = f"""SELECT id,codigo_item,quantidade,motivo,observacao  
                         FROM sistema_epi.tb_solicitacoes
                         WHERE id_solicitacao = '{id_solicitante}'
                         ORDER BY id asc;"""

    cur.execute(query)
    info_gerais = cur.fetchall()

    cur.execute(query_solicitacoes)
    equipamento = cur.fetchall()

    dados = {
        'id_solicitacao': info_gerais[-1][1],
        'matricula': info_gerais[-1][2],
        'funcionario': info_gerais[-1][7],
        'data_solicitacao': info_gerais[0][8],
        'data_assinado': info_gerais[-1][11],
    }

    equipamentos = []
    
    for row in equipamento:
        observacao = row[4] if row[4] is not None else ""

        equipamentos.append({
            'id': row[0],
            'codigo': row[1],
            'quantidade': row[2],
            'motivo': row[3],
            'observacao':observacao
        })

    dados['equipamentos'] = equipamentos

    print(dados)

    # Exemplo de resposta de volta para o cliente
    return jsonify(dados)

@app.route('/timeline', methods=['POST'])
@login_required
def timeline_os():

    dados = request.get_json()

    id_solicitacao = dados['id_solicitacao']
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = f"""SELECT * 
            FROM sistema_epi.tb_historico_solicitacoes
            WHERE id_solicitacao = '{id_solicitacao}'"""
    
    df_timeline = pd.read_sql_query(query, conn)

    print(df_timeline)

    df_timeline = df_timeline.sort_values(by='id', ascending=True)

    df_timeline = df_timeline.values.tolist()

    return jsonify (id_solicitacao,df_timeline)

def gerar_id_solicitacao():
    
    """
    Função para gerar id único
    """

    return str(uuid.uuid1())

def query_filtro_historico():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)


    query_funcionarios = """SELECT DISTINCT funcionario_completo
                                    FROM (
                                        SELECT CONCAT(s.funcionario_recebe, ' - ', f.nome) AS funcionario_completo, s.*
                                        FROM sistema_epi.tb_solicitacoes s
                                        JOIN requisicao.funcionarios f ON s.funcionario_recebe = f.matricula
                                    ) AS subconsulta
                              """
    
    query_solicitante = """SELECT DISTINCT solicitante_completo
                                FROM (
                                    SELECT CONCAT(s.matricula_solicitante, ' - ', f.nome) AS solicitante_completo, s.*
                                    FROM sistema_epi.tb_solicitacoes s
                                    JOIN requisicao.funcionarios f ON s.matricula_solicitante = f.matricula
                                ) AS subconsulta
                            """
    
    query_itens = """
                        SELECT DISTINCT CONCAT(codigo,' - ',descricao)
                        FROM sistema_epi.tb_itens
                        """
    
    cur.execute(query_funcionarios)
        
    funcionarios = cur.fetchall()

    cur.execute(query_solicitante)
    
    solicitantes = cur.fetchall()
    
    cur.execute(query_itens)
    
    itens = cur.fetchall()

    return funcionarios, solicitantes, itens

def cardsPaginaInicial(cur):

    query_quantidade_assinatura = """SELECT COUNT(*) AS quantidade_assinados
                                FROM sistema_epi.tb_assinatura;"""
    cur.execute(query_quantidade_assinatura)
    quantidade_assinados = cur.fetchone()[0]

    query_quantidade_devolucao = """SELECT COUNT(*) AS quantidade_devolucao
                                    FROM sistema_epi.tb_solicitacoes
                                    WHERE status_devolucao IS NOT NULL;"""
    cur.execute(query_quantidade_devolucao)
    quantidade_devolucao = cur.fetchone()[0]

    query_quantidade_trabalhadores = """SELECT COUNT(DISTINCT funcionario_recebe) AS quantidade_trabalhadores
                                        FROM sistema_epi.tb_solicitacoes;"""
    cur.execute(query_quantidade_trabalhadores)
    quantidade_trabalhadores = cur.fetchone()[0]

    query_quantidade_solicitacoes = """SELECT COUNT(id_solicitacao) AS quantidade_solicitacoes
                                    FROM sistema_epi.tb_solicitacoes;"""
    cur.execute(query_quantidade_solicitacoes)
    quantidade_solicitacoes = cur.fetchone()[0]

    return quantidade_assinados,quantidade_devolucao,quantidade_trabalhadores,quantidade_solicitacoes

def verifica_existencia_solicitacao(cur, matricula_recebedor, codigo):
    """
    Verifica se já existe uma solicitação com o mesmo matricula_recebedor e código na tabela sistema_epi.tb_solicitacoes.
    Retorna True se existir, False caso contrário.
    """
    sql ="""SELECT *
            FROM sistema_epi.tb_solicitacoes
        WHERE funcionario_recebe = %s AND codigo_item = %s AND status_devolucao IS NULL
        ORDER BY id_solicitacao DESC
        LIMIT 1;"""

    cur.execute(sql,(matricula_recebedor,codigo))

    return cur.fetchone() is not None

def verifica_existencia_data_devolucao(cur, matricula_recebedor, codigo):
    """
    Verifica se já existe uma solicitação com o mesmo matricula_recebedor e código na tabela sistema_epi.tb_solicitacoes.
    Retorna True se existir, False caso contrário.
    """
    sql ="""SELECT *
            FROM sistema_epi.tb_solicitacoes
        WHERE funcionario_recebe = %s AND codigo_item = %s AND status_devolucao NOTNULL
        ORDER BY id_solicitacao DESC
        LIMIT 1;"""
    
    print(sql)

    cur.execute(sql,(matricula_recebedor,codigo))

    return cur.fetchone() is not None

def input_tb_solicitacoes(campos_solicitacao,id_solicitacao):
    """
    Função para inputar dados na tabela de solicitações do schema: sistema_epi
    """

    # Dicionário de novas chaves e valores
    novos_campos = {'id_solicitacao': id_solicitacao, 'status': 'Aguardando Assinatura'}

    # Iterar sobre cada dicionário na lista
    for item in campos_solicitacao:
        # Adicionar as novas chaves e valores
        item.update(novos_campos)
    
    print("campos_solicitacao",campos_solicitacao)

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                    password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Iterar sobre os dados e inserir no banco de dados
    for item in campos_solicitacao:
        id_solicitacao = list(item.values())[6]
        matricula_solicitante = list(item.values())[0].split()[0]
        codigo = list(item.values())[1]
        quantidade = list(item.values())[2]
        motivo = list(item.values())[5]
        observacao = list(item.values())[4]
        matricula_recebedor = list(item.values())[3].split()[0]

        print(f"id_solicitacao:{id_solicitacao},matricula_solicitante:{matricula_solicitante},codigo:{codigo},quantidade:{quantidade},motivo:{motivo},observacao:{observacao},matricula_recebedor:{matricula_recebedor}")

        if verifica_existencia_solicitacao(cur, matricula_recebedor,codigo) == True:

            query ="""UPDATE sistema_epi.tb_solicitacoes
            SET status_devolucao = %s
            WHERE id = (
                SELECT id
                    FROM sistema_epi.tb_solicitacoes
                WHERE funcionario_recebe = %s AND codigo_item = %s AND status_devolucao IS NULL
                ORDER BY id_solicitacao DESC
                LIMIT 1);"""
            
            cur.execute(query, (motivo,matricula_recebedor,codigo))

        sql = """INSERT INTO sistema_epi.tb_solicitacoes 
            (id_solicitacao, matricula_solicitante, codigo_item, quantidade, motivo,funcionario_recebe,observacao)
        VALUES
            (%s, %s, %s, %s, %s, %s, %s);"""

        cur.execute(sql,(id_solicitacao,matricula_solicitante,codigo,quantidade,motivo,matricula_recebedor,observacao))

    conn.commit()

def base_tb_historico(id_solicitacao,status_assinatura,motivo_assinatura):
    """
    Função para inputar dados na tabela de histórico do schema: sistema_epi
    A cada atualização será criado um registro
    """
    
    # campos_solicitacao = [{}]

    # Dicionário de novas chaves e valores
    # campos_solicitacao = {'id_solicitacao': id_solicitacao, 'status': 'Aguardando Assinatura'}

    # Iterar sobre cada dicionário na lista
    # for item in campos_solicitacao:
    #     # Adicionar as novas chaves e valores
    #     item.update(novos_campos)

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                    password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # Iterar sobre os dados e inserir no banco de dados
    # for item in campos_solicitacao:
    status = status_assinatura
    motivo = motivo_assinatura

    print(id_solicitacao, status)

    sql = """INSERT INTO sistema_epi.tb_historico_solicitacoes 
        (id_solicitacao, status,motivo)
    VALUES
        (%s, %s,%s);"""

    cur.execute(sql,(id_solicitacao,status,motivo))

    conn.commit()

# Operadores
@app.route('/operadores', methods=['GET'])
def listar_operadores():
    """
    Função para buscar operadores disponíveis
    """
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    sql = "select concat(matricula, ' - ', nome) from requisicao.funcionarios where ativo = TRUE"
    cur.execute(sql)

    data = cur.fetchall()

    return jsonify(data)

# Itens
@app.route('/itens', methods=['GET'])
def listar_itens():
    """
    Função para buscar itens
    """
    
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    sql = "select concat(codigo, ' - ', descricao) from sistema_epi.tb_itens"
    cur.execute(sql)

    data = cur.fetchall()

    return jsonify(data)

# Setor do operador
@app.route('/setor-operador/<operador>', methods=['GET'])
def setor_operador(operador):
    """
    Função para buscar setor do operador
    """

    operador = operador.split()[0]

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    sql = "select setor from requisicao.funcionarios where matricula = %s"
    cur.execute(sql,(operador,))

    setor = cur.fetchone()

    return jsonify(setor[0])

# Formulário de solicitação
@app.route('/solicitacao-material', methods=['GET','POST'])
def rota_solicitacao_material():

    """
    Rota para de solicitação de material
    """

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    usuario = session.get('user_id')

    sql = f"select concat(matricula, ' - ', nome) from requisicao.funcionarios where matricula = '{usuario}'"
    cur.execute(sql)

    data = cur.fetchall()

    solicitante = data[0][0]

    nome = solicitante.split(' - ')[1]

    # Renderize o template e passe o parâmetro de sucesso, se aplicável
    return render_template('solicitacao-material.html', solicitante=solicitante,nome=nome)

# Dashboard
@app.route('/dashboard', methods=['GET','POST'])
def dashboard():
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    # sql = "select * from sistema_epi.tb_solicitacoes ts left join sistema_epi.tb_assinatura ta on ts.id_solicitacao = ta.id_solicitacao where data_assinatura is null"
    sql = """SELECT
            solic.id,
            solic.id_solicitacao,
            solic.matricula_solicitante,
            solic.codigo_item,
            solic.quantidade,
            solic.motivo,   
            solic.setor_solicitante,   
            solic.funcionario_recebe,   
            solic.data_solicitada,   
            solic.status_devolucao,   
            hist.status,
            ass.assinatura,
            ass.data_assinatura
        FROM sistema_epi.tb_solicitacoes AS solic
        LEFT JOIN (
            SELECT
                id_solicitacao,
                status,
                ROW_NUMBER() OVER (PARTITION BY id_solicitacao ORDER BY data_modificacao DESC) AS row_num
            FROM sistema_epi.tb_historico_solicitacoes
        ) AS hist ON hist.id_solicitacao = solic.id_solicitacao AND hist.row_num = 1
        LEFT JOIN sistema_epi.tb_assinatura AS ass ON ass.id_solicitacao = solic.id_solicitacao
        WHERE solic.status_devolucao IS null and data_assinatura is null ORDER BY solic.data_solicitada DESC"""

    tb_solicitacoes = pd.read_sql_query(sql, conn)

    tb_solicitacoes['data_solicitada'] = pd.to_datetime(tb_solicitacoes['data_solicitada'])
    tb_solicitacoes['data_solicitada_iso'] = tb_solicitacoes['data_solicitada'].dt.strftime("%Y-%m-%dT%H:%M:%S")
    tb_solicitacoes['data_solicitada'] = tb_solicitacoes['data_solicitada'].dt.strftime('%d/%m/%Y')

    sql = f"select concat(matricula, ' - ', nome) from requisicao.funcionarios"
    
    tb_sql = pd.read_sql_query(sql, conn)

    mapeamento = dict(zip(tb_sql['concat'].str.extract('(\d+)', expand=False), tb_sql['concat']))

    # Substituir na coluna matricula_solicitante
    tb_solicitacoes['matricula_solicitante'] = tb_solicitacoes['matricula_solicitante'].astype(str)
    tb_solicitacoes['matricula_solicitante'] = tb_solicitacoes['matricula_solicitante'].replace(mapeamento)

    tb_solicitacoes['funcionario_recebe'] = tb_solicitacoes['funcionario_recebe'].astype(str)
    tb_solicitacoes['funcionario_recebe'] = tb_solicitacoes['funcionario_recebe'].replace(mapeamento)


    tb_solicitacoes_list = tb_solicitacoes.values.tolist()

    # print(data)
    # Renderize o template e passe o parâmetro de sucesso, se aplicável
    return render_template('dashboard.html',data=tb_solicitacoes_list)

@app.route('/atualizar-dashboard')
def atualizar_dashboard():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)

    sql = """SELECT
            solic.id,
            solic.id_solicitacao,
            solic.matricula_solicitante,
            solic.codigo_item,
            solic.quantidade,
            solic.motivo,   
            solic.setor_solicitante,   
            solic.funcionario_recebe,   
            solic.data_solicitada,   
            solic.status_devolucao,   
            hist.status,
            ass.assinatura,
            ass.data_assinatura
        FROM sistema_epi.tb_solicitacoes AS solic
        LEFT JOIN (
            SELECT
                id_solicitacao,
                status,
                ROW_NUMBER() OVER (PARTITION BY id_solicitacao ORDER BY data_modificacao DESC) AS row_num
            FROM sistema_epi.tb_historico_solicitacoes
        ) AS hist ON hist.id_solicitacao = solic.id_solicitacao AND hist.row_num = 1
        LEFT JOIN sistema_epi.tb_assinatura AS ass ON ass.id_solicitacao = solic.id_solicitacao
        WHERE solic.status_devolucao IS null and data_assinatura is null ORDER BY solic.data_solicitada DESC"""

    tb_solicitacoes = pd.read_sql_query(sql, conn)

    tb_solicitacoes['data_solicitada'] = pd.to_datetime(tb_solicitacoes['data_solicitada'])
    tb_solicitacoes['data_solicitada_iso'] = tb_solicitacoes['data_solicitada'].dt.strftime("%Y-%m-%dT%H:%M:%S")
    tb_solicitacoes['data_solicitada'] = tb_solicitacoes['data_solicitada'].dt.strftime('%d/%m/%Y')

    sql = f"select concat(matricula, ' - ', nome) from requisicao.funcionarios"
    
    tb_sql = pd.read_sql_query(sql, conn)

    mapeamento = dict(zip(tb_sql['concat'].str.extract('(\d+)', expand=False), tb_sql['concat']))

    # Substituir na coluna matricula_solicitante
    tb_solicitacoes['matricula_solicitante'] = tb_solicitacoes['matricula_solicitante'].astype(str)
    tb_solicitacoes['matricula_solicitante'] = tb_solicitacoes['matricula_solicitante'].replace(mapeamento)

    tb_solicitacoes['funcionario_recebe'] = tb_solicitacoes['funcionario_recebe'].astype(str)
    tb_solicitacoes['funcionario_recebe'] = tb_solicitacoes['funcionario_recebe'].replace(mapeamento)


    tb_solicitacoes_list = tb_solicitacoes.values.tolist()

    return jsonify({
        "solicitacoes": tb_solicitacoes_list
    })

@app.route('/solicitacao', methods=['POST'])
def criar_solicitacao():
    """
    Função para receber as solicitações que foram geradas
    """

    dados = request.get_json()
    print(dados)

    # Dicionário para armazenar dados agrupados pelo inputOperador
    dados_agrupados = {}

    # Iterar sobre a resposta original
    for item in dados:
        
        # Obter o valor do inputOperador usando o índice
        input_operador = list(item.values())[3]

        # Verificar se já existe uma entrada para esse inputOperador no dicionário dadosAgrupados
        if input_operador not in dados_agrupados:
            # Se não existir, crie uma nova entrada com uma lista vazia
            dados_agrupados[input_operador] = []

        # Adicione o item à lista correspondente ao inputOperador
        dados_agrupados[input_operador].append(item)

    # Crie uma lista para cada valor único de inputOperador
    listas = {}
    motivo = "Primeira Assinatura"
    status = "Aguardando Assinatura"

    for input_operador, lista in dados_agrupados.items():
        listas[input_operador] = lista

    lista_id_solicitacao = []

    # print(listas.items())

    # Imprima as listas
    for input_operador, lista in listas.items():
        id_solicitacao = gerar_id_solicitacao()
        print(id_solicitacao)
        lista_id_solicitacao.append(id_solicitacao)
        input_tb_solicitacoes(lista,id_solicitacao)
    
    for id in lista_id_solicitacao:
        print(id)
        base_tb_historico(id,status,motivo)
    
    return redirect(url_for('rota_solicitacao_material'))

@app.route('/vida-util', methods=['POST'])
def buscar_vida_util():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    dados = request.get_json()
    cur.execute('select vida_util from sistema_epi.tb_itens')
    vida_util = cur.fetchone()

    return jsonify(vida_util)

@app.route('/alterar-dados', methods=['POST'])
def alterar_dados():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        # Recebe os dados do corpo da solicitação como JSON
        dados = request.get_json()

        # Acesse os dados específicos, por exemplo, os equipamentos
        equipamentos = dados.get('equipamentos')
        id_solicitacao = equipamentos[0]['id_solicitacao']
        print(id_solicitacao)

        query_assinatura = f"""DELETE
                FROM sistema_epi.tb_assinatura
                WHERE id_solicitacao = '{id_solicitacao}'"""
        
        cur.execute(query_assinatura)

        motivo = "Editou o Item"
        status = "Aguardando Assinatura"

        base_tb_historico(id_solicitacao,status,motivo)

        # Faça algo com os dados, como salvá-los no banco de dados
        for equipamento in equipamentos:
            idExecucao = equipamento.get('idExecucao')
            equipamento_nome = equipamento.get('equipamento')
            quantidade = equipamento.get('quantidade')
            motivo = equipamento.get('motivo')
            observacao = equipamento.get('observacao')

            cur.execute(""" UPDATE sistema_epi.tb_solicitacoes
                    SET codigo_item=%s, quantidade=%s, motivo=%s, observacao=%s
                    WHERE id = %s
                    """, (equipamento_nome, quantidade, motivo, observacao, idExecucao))
            
            print(idExecucao,equipamento_nome,quantidade,motivo)

            conn.commit()
            
        conn.close()

            # Responda ao cliente
        return jsonify({'mensagem': 'Dados recebidos com sucesso!'})

    except Exception as e:
        return jsonify({'erro': str(e)}), 500

@app.route('/salvar-novos-dados',methods=['POST'])
def salvar_dados():
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    dados = request.get_json()

    id_solicitacao = dados['id_solicitacao']
    equipamento = dados['equipamento']
    quantidade = dados['quantidade']
    motivo = dados['motivo']
    matricula_nome_solicitante = dados['matricula_solicitante']
    matricula_nome_recebedor = dados['matricula_recebedor']
    observacao = dados['observacao']

    matricula_solicitante,nome_solicitante = matricula_nome_solicitante.split(" - ")
    matricula_recebedor,nome_recebedor = matricula_nome_recebedor.split(" - ")

    sql = """INSERT INTO sistema_epi.tb_solicitacoes 
            (id_solicitacao, matricula_solicitante, codigo_item, quantidade, motivo,funcionario_recebe,observacao)
        VALUES
            (%s, %s, %s, %s, %s, %s,%s);"""

    cur.execute(sql,(id_solicitacao,matricula_solicitante,equipamento,quantidade,motivo,matricula_recebedor,observacao))

    query_assinatura = f"""DELETE FROM sistema_epi.tb_assinatura
                        WHERE id_solicitacao = '{id_solicitacao}'
                        """
    
    cur.execute(query_assinatura)

    codigo,nome_equipamento = equipamento.split(" - ")

    motivo = "Adicionou o Item - " + codigo
    status = "Aguardando Assinatura"

    conn.commit()
    conn.close()

    base_tb_historico(id_solicitacao,status,motivo)

    conn.close()

    return jsonify("Sucesso")

@app.route('/add-item-padrao',methods=['POST'])
def add_item_padraoo():
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    dados = request.get_json()

    equipamento_adicionado = dados['equipamento_adicionado']
    quantidade_adicionado = dados['quantidade_adicionado']
    motivo_adicionado = dados['motivo_adicionado']
    funcionario_adicionado = dados['funcionario_adicionado']
    observacao_adicionado = dados['observacao_adicionado']
    nome_padrao_adicionado = dados['nome_padrao_adicionado']
    solicitante_adicionado = dados['solicitante_adicionado']

    print(equipamento_adicionado,quantidade_adicionado,motivo_adicionado,funcionario_adicionado,observacao_adicionado,nome_padrao_adicionado,solicitante_adicionado)

    query = """ INSERT INTO sistema_epi.padrao_solicitacao (matricula_solicitante,nome,codigo_item,motivo,quantidade,funcionario_recebe,observacao)
                VALUES (%s,%s,%s,%s,%s,%s,%s) """
    
    values = (solicitante_adicionado,nome_padrao_adicionado,equipamento_adicionado,motivo_adicionado,quantidade_adicionado,funcionario_adicionado,observacao_adicionado)

    cur.execute(query,values)

    conn.commit()

    cur.close()
    conn.close()

    return jsonify("Sucesso")

@app.route('/excluir-solicitacao', methods=['POST'])
@login_required
def excluir_solicitacao():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    data = request.get_json()

    id_solicitacao = data['id_solicitacao']

    query_solicitacao = f"""DELETE FROM sistema_epi.tb_solicitacoes
            WHERE id_solicitacao = '{id_solicitacao}';
            """
    
    cur.execute(query_solicitacao)

    query_historico = f"""DELETE FROM sistema_epi.tb_historico_solicitacoes
            WHERE id_solicitacao = '{id_solicitacao}';
            """
    
    cur.execute(query_historico)

    query_assinatura = f"""DELETE FROM sistema_epi.tb_assinatura
            WHERE id_solicitacao = '{id_solicitacao}';
            """
    
    cur.execute(query_assinatura)

    conn.commit()
    conn.close()

    return 'Dados recebidos com sucesso!'

@app.route('/excluir-equipamento', methods=['POST'])
@login_required
def excluir_equipamento():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    data = request.get_json()

    id = data['id']
    id_solicitante = data['id_solicitante']
    codigo = data['codigo']

    query_assinatura = f"""DELETE FROM sistema_epi.tb_assinatura
                        WHERE id_solicitacao = '{id_solicitante}'
                        """
    
    cur.execute(query_assinatura)

    query_solicitacao = f"""DELETE FROM sistema_epi.tb_solicitacoes WHERE id = {id}"""
    print(query_solicitacao)

    cur.execute(query_solicitacao)

    motivo = "Excluiu Item - " + codigo
    status = "Aguardando Assinatura"

    conn.commit()
    conn.close()

    base_tb_historico(id_solicitante,status,motivo)

    conn.close()

    return 'Dados recebidos com sucesso!'

@app.route('/excluir-assinatura', methods=['POST'])
@login_required
def excluir_assinatura():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER, password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    data = request.get_json()

    id_solicitante = data['id_solicitante']

    query_assinatura = f"""DELETE FROM sistema_epi.tb_assinatura
                        WHERE id_solicitacao = '{id_solicitante}'
                        """
    
    cur.execute(query_assinatura)

    conn.commit()
    conn.close()

    return jsonify('Dados recebidos com sucesso!')


def condicao_historico(matricula, data_solicit,solicitante_historico,equipamento_historico):

    consulta_padrao = """
                    SELECT 
                        f_solicitante.nome as nome_solicitante,
                        s.codigo_item,
                        s.quantidade,
                        s.motivo,
                        f.nome as nome_funcionario_recebe,
                        s.data_solicitada
                    FROM 
                        sistema_epi.tb_solicitacoes s
                    JOIN 
                        requisicao.funcionarios f_solicitante ON s.matricula_solicitante = f_solicitante.matricula
                    JOIN 
                        requisicao.funcionarios f ON s.funcionario_recebe = f.matricula
                    JOIN 
                        sistema_epi.tb_assinatura a ON s.id_solicitacao = a.id_solicitacao
                    WHERE 1=1"""
    
    if matricula:
        consulta_padrao += f" AND s.funcionario_recebe = '{matricula}'"
    if data_solicit:
        mes_inicial, mes_final = data_solicit.split(' - ')
        mes_inicial = datetime.strptime(mes_inicial, '%d/%m/%Y').date()
        mes_final = datetime.strptime(mes_final, '%d/%m/%Y').date() + timedelta(days=1)

        mes_inicial_formatado = mes_inicial.strftime('%Y-%m-%d')
        mes_final_formatado = mes_final.strftime('%Y-%m-%d')
        consulta_padrao += f" AND s.data_solicitada >= '{mes_inicial_formatado}' AND s.data_solicitada <= '{mes_final_formatado}'"
    if solicitante_historico:
        consulta_padrao += f" AND s.matricula_solicitante = '{solicitante_historico}'"
    if equipamento_historico:
        consulta_padrao += f" AND s.codigo_item = '{equipamento_historico}'"

    consulta_padrao += ' ORDER BY s.id desc;'

    return consulta_padrao

@app.route('/historico', methods=['GET','POST'])
@login_required
def historico():

    if request.method == 'POST':

        conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        data = request.get_json()

        matricula = data['matricula']

        data_solicit = data['data']

        solicitante_historico = data['solicitante_historico']

        equipamento_historico = data['equipamento_historico']

        query_historico = condicao_historico(matricula,data_solicit,solicitante_historico,equipamento_historico)

        print(query_historico)
        
        cur.execute(query_historico)
        
        historicos = cur.fetchall()

        # Transformar a lista de listas em uma lista de dicionários
        historicos_dicts = []
        for historico in historicos:
            historico_dict = {
                'Solicitante': historico[0],
                'Equipamento': historico[1],
                'Quantidade': historico[2],
                'Motivo': historico[3],
                'Funcionario': historico[4],
                'DataSolicitada': historico[5].strftime('%d/%m/%Y')  # Convertendo para uma string de data formatada
            }
            historicos_dicts.append(historico_dict)

        return jsonify(historicos_dicts)

    else:
        funcionarios, solicitantes, itens = query_filtro_historico()
    
    return render_template('historico.html',funcionarios=funcionarios,solicitantes=solicitantes,itens=itens)
    
@app.route('/ficha', methods=['GET','POST'])
@login_required
def ficha():

    if request.method == 'POST':

        conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
        cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

        data = request.get_json()

        operadorFicha = data['operadorFicha']

        matricula, operador = operadorFicha.split(' - ')

        query_funcionario = f""" SELECT funcao, data_admissao,setor FROM requisicao.funcionarios WHERE nome = '{operador}'"""
        
        cur.execute(query_funcionario)
        funcao_admissao = cur.fetchall()

        dateFicha = data['dateFicha']

        print(dateFicha)

        query = f"""
                    SELECT 
                        a.data_assinatura,
                        s.quantidade,
                        s.codigo_item,
                        itens.ca,
                        s.status_devolucao,
                        s.funcionario_recebe,
                        s.data_devolucao,
                        a.assinatura
                    FROM 
                    sistema_epi.tb_solicitacoes s
                    JOIN 
                    sistema_epi.tb_assinatura a
                    ON 
                    s.id_solicitacao = a.id_solicitacao
                    LEFT JOIN sistema_epi.tb_itens itens ON s.codigo_item = CONCAT(itens.codigo || ' - ' || itens.descricao)
                    WHERE 1=1 AND s.funcionario_recebe = '{matricula}'
                """

        if dateFicha:
            mes_inicial, mes_final = dateFicha.split(' - ')
            mes_inicial = datetime.strptime(mes_inicial, '%d/%m/%Y').date()
            mes_final = datetime.strptime(mes_final, '%d/%m/%Y').date() + timedelta(days=1)

            mes_inicial_formatado = mes_inicial.strftime('%Y-%m-%d')
            mes_final_formatado = mes_final.strftime('%Y-%m-%d')
            query += f" AND a.data_assinatura >= '{mes_inicial_formatado}' AND a.data_assinatura <= '{mes_final_formatado}'"

        cur.execute(query)
        lista_solicitacoes = cur.fetchall()

        num_solicitacoes = len(lista_solicitacoes)

        if num_solicitacoes == 0:
            return jsonify('Vazio')

        for i in range(num_solicitacoes):
            assinatura = lista_solicitacoes[i][7]
            del(lista_solicitacoes[i][7])
            assinatura_bytes = bytes(assinatura)
            lista_solicitacoes[i].append(assinatura_bytes)

        wb = load_workbook('FICHA DE EPI Atualizada 15.01.2024.xlsx')

        ws = wb.active

        if num_solicitacoes != 0:
            num_solicitacoes = len(lista_solicitacoes)


        # Aumentar a quantidade de Linhas
        for i in range(num_solicitacoes):
            # Define a linha de destino
            linha_destino = 27 + i

            # Copia o valor e o estilo da linha 27 para a linha de destino
            for coluna in range(1, 9):  # A coluna 1 é a A, a coluna 2 é a B, etc.
                ws.cell(row=linha_destino, column=coluna).font = copy.copy(ws.cell(row=27, column=coluna).font)
                ws.cell(row=linha_destino, column=coluna).fill = copy.copy(ws.cell(row=27, column=coluna).fill)
                ws.cell(row=linha_destino, column=coluna).border = copy.copy(ws.cell(row=27, column=coluna).border)
                ws.cell(row=linha_destino, column=coluna).alignment = copy.copy(ws.cell(row=27, column=coluna).alignment)
                        # Access and copy the line height from row 27
                line_height = ws.row_dimensions[27].height  # Access height from source row
                ws.row_dimensions[linha_destino].height = line_height  # Set the same height for the target row

        num_assinaturas = []
        # Inserir a tabela do SQL 
        for i in range(num_solicitacoes):
            # Define a linha de destino
            num_assinaturas.append(i)
            ws['B4'] = operador
            ws['B5'] = int(matricula.replace(',', ''))
            ws['B6'] = funcao_admissao[0][0]
            ws['B7'] = funcao_admissao[0][1]
            ws['B8'] = funcao_admissao[0][2]

            linha_destino = 27 + i

            data_formatada = lista_solicitacoes[i][0].strftime("%d/%m/%Y")
            ws.cell(row=linha_destino, column=1).value = data_formatada
            ws.cell(row=linha_destino, column=2).value = lista_solicitacoes[i][1]  # Quantidade
            ws.cell(row=linha_destino, column=3).value = lista_solicitacoes[i][2]  # Código do item
            ws.cell(row=linha_destino, column=4).value = lista_solicitacoes[i][3]  # CA do item
            ws.cell(row=linha_destino, column=5).value = lista_solicitacoes[i][6]  # Devolução do item
            ws.cell(row=linha_destino, column=6).value = lista_solicitacoes[i][4]  # Motivo
            # assinatura = lista_solicitacoes[i][5]
            base64_string = lista_solicitacoes[i][7]

            # Remove o prefixo 'data:image/png;base64,' para obter apenas a parte codificada em base64
            image_data = base64_string.split(b';base64,')[1]

            # Decodifica a string base64
            image_bytes = base64.b64decode(image_data)

            # Cria um objeto BytesIO para criar uma imagem PIL a partir dos bytes
            image_buffer = BytesIO(image_bytes)

            # Abre a imagem PIL
            image = Image.open(image_buffer)

            image_fixo = Image.open(image_buffer)

            image_fixo.save(f"assinaturaL.png")

            img_fixo = imge(f"assinaturaL.png")

            img_fixo.height = 130
            img_fixo.width = 150
            
            image.save(f"assinatura{i}.png")

            img = imge(f"assinatura{i}.png")
            img.height = 130
            img.width = 150
            ws.add_image(img, f'G{linha_destino}')

        ws.add_image(img_fixo,'B22')
        
        wb.save(r'downloads/Nova_ficha.xlsx')
        
        wb.close()

        for num in num_assinaturas:
            os.remove(f"assinatura{num}.png")
        os.remove(f"assinaturaL.png")

        return jsonify('OK')
    
    else:

        funcionarios, solicitantes,itens = query_filtro_historico()
    
    return render_template('ficha.html',funcionarios=funcionarios)

@app.route('/downloads/Nova_ficha.xlsx', methods=['GET'])
def download_modelo_excel():
    # Caminho para o arquivo modelo CSV
    excel_filename = r'downloads/Nova_ficha.xlsx'

    # Envie o arquivo para download
    return send_file(excel_filename, as_attachment=True, mimetype='text/csv')

# AINDA NÃO UTLIZADA, POREM VAI AJUDAR NA DOCUMENTAÇÃO
@app.route('/pegar-assinatura', methods=['GET'])
@login_required
def pegar_assinatur():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    consulta = "SELECT id_solicitacao, assinatura, data_assinatura FROM sistema_epi.tb_assinatura"
    cur.execute(consulta)
    resultados = cur.fetchall()

    # 2. Extrair os dados da consulta
    dados_assinatura = []
    for resultado in resultados:
        id_solicitacao, assinatura_memoryview, data_assinatura = resultado
        
        # Converter memoryview para bytes e, em seguida, decodificar
        assinatura_bytes = bytes(assinatura_memoryview)
        assinatura_legivel = assinatura_bytes.decode('utf-8')  # ou outro método apropriado

        # Adicione os dados a uma lista
        dados_assinatura.append({
            'id_solicitacao': id_solicitacao,
            'assinatura': assinatura_legivel,
            'data_assinatura': data_assinatura
        })

    return 'sucess'

# Funcionários

#Listar setores para funcionários
@app.route('/listar-setores-funcionarios', methods=['GET'])
def lista_setor_cadastro():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query = """select distinct setor from requisicao.funcionarios f"""

    cur.execute(query)
    lista_setores = cur.fetchall()

    return lista_setores

#Receber cadastro de funcionario
@app.route('/cadastrar-funcionario', methods=['POST'])
def cadastrar_funcionario():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    nome = request.form.get('nome')
    matricula = request.form.get('matricula')
    setor = request.form.get('setor')
    data_admissao = request.form.get('data_admissao')

    print(data_admissao)

    query_consulta = """select * from requisicao.funcionarios where matricula = %s"""
    cur.execute(query_consulta,(matricula,))

    data = cur.fetchall()

    if len(data) > 0:
        tipo_mensagem = 'aviso'
        mensagem = 'Matrícula ja existe'

    else:
        query_insert = """insert into requisicao.funcionarios (matricula,nome,data_admissao,setor) values(%s,%s,%s,%s)"""
        cur.execute(query_insert,(matricula,nome,data_admissao,setor,))

        conn.commit()

        tipo_mensagem = 'sucesso'
        mensagem = 'Cadastrado com sucesso'

    conn.close()

    return jsonify({
        'tipo_mensagem':tipo_mensagem,
        'mensagem':mensagem,
    })

# Padrão
@app.route('/salvar-novo-padrao', methods=['POST'])
def salvar_novo_padrao():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    dados = request.get_json()

    matricula = dados['dados'][0]['solicitante'].split()[0]
    nome_padrao = dados['nome_padrao']

    for item in dados['dados']:
        codigo_item=item['inputCodigo']
        quantidade=item['inputQuantidade']
        funcionario_recebe=item['inputOperador']
        motivo=item['radioSubstituicao']
        observacao=item['observacaoSolicitacao']

        query_add = """insert into sistema_epi.padrao_solicitacao (matricula_solicitante,nome,codigo_item,motivo,quantidade,funcionario_recebe,observacao) values (%s,%s,%s,%s,%s,%s,%s)"""

        cur.execute(query_add,(matricula,nome_padrao,codigo_item,motivo,quantidade,funcionario_recebe,observacao))
    
    conn.commit()

    return 'sucess'

@app.route('/verificar-nome-padrao', methods=['POST'])
def verificar_nome_padrao():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    dados = request.get_json()

    matricula = dados['nome_solicitante'].split()[0]
    nome_padrao = dados['nome_padrao']

    query_verificar = """select * from sistema_epi.padrao_solicitacao where matricula_solicitante = %s and nome = %s"""
    cur.execute(query_verificar,(matricula,nome_padrao))
    data_verificar = cur.fetchall()

    if len(data_verificar) > 0:
        response = {'tipo':'aviso','mensagem': 'Nome já existe, escolha outro'}
    else:
        response = {'tipo':'success','mensagem': 'Nome disponível. Salvo!'}

    cur.close()
    conn.close()
    
    return jsonify(response)

@app.route('/buscar-padroes', methods=['GET'])
def buscar_padrao():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    user_login = session['user_id']

    cur.execute("select distinct nome from sistema_epi.padrao_solicitacao where matricula_solicitante = %s",(user_login,))

    padroes=cur.fetchall()

    return jsonify({'padroes':padroes})

@app.route('/popular-padroes', methods=['POST'])
def popular_padrao():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    data = request.get_json()
    print(data)
    user_login = session['user_id']

    cur.execute("select * from sistema_epi.padrao_solicitacao where matricula_solicitante = %s and nome = %s",(user_login,data['nome_padrao']))

    itens=cur.fetchall()
    itens_dict = [dict(item) for item in itens]

    return jsonify({'itens':itens_dict})

@app.route('/excluir-item-padrao', methods=['POST'])
def excluir_item_padrao():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    data = request.get_json()  # Recebe os dados da requisição
    nome_padrao = data.get('nome_padrao')
    codigo_item = data.get('codigo_item')
    matricula_solicitante = data.get('matricula_solicitante')
    funcionario_recebe = data.get('funcionario_recebe')

    print(nome_padrao,codigo_item,matricula_solicitante,funcionario_recebe)

    delete_query = F"""DELETE
                    FROM sistema_epi.padrao_solicitacao
                    WHERE matricula_solicitante = '{matricula_solicitante}' AND nome = '{nome_padrao}' 
                    AND codigo_item = '{codigo_item}' AND funcionario_recebe = '{funcionario_recebe}'"""
    
    cur.execute(delete_query)

    conn.commit()
    conn.close()

    return jsonify('success')

@app.route('/crud-equipamento', methods=['POST'])
def crud_equipamento():

    data = request.get_json()

    print(data)

    codigo = data.get('codigo')
    nome_equipamento = data.get('nome')
    vida_util = data.get('vida_util')
    ca = data.get('ca')
    acao = data.get('acao')  # Ex: 'create', 'update', 'delete'

    if ca == '':
        ca = None

    print(codigo,nome_equipamento,vida_util,ca,acao)

    if acao == 'create':
        crud.create_equipamento(codigo, nome_equipamento, vida_util, ca)
        return jsonify({'message': 'Equipamento criado com sucesso'}), 201
    elif acao == 'update':
        crud.update_equipamento(codigo, vida_util, ca)
        return jsonify({'message': 'Equipamento atualizado com sucesso'}), 200
    elif acao == 'delete':
        crud.delete_equipamento(codigo)
        return jsonify({'message': 'Equipamento deletado com sucesso'}), 200
    else:
        return jsonify({'message': 'Ação inválida'}), 400

@app.route('/equipamentos', methods=['GET'])
def equipamentos():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query_equipamentos = 'SELECT * FROM sistema_epi.tb_itens'

    cur.execute(query_equipamentos)
    equipamentos = cur.fetchall()

    return render_template("equipamentos.html",equipamentos=equipamentos)

@app.route('/funcionarios', methods=['GET'])
def funcionarios():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    query_funcionarios = 'SELECT * FROM requisicao.funcionarios'

    cur.execute(query_funcionarios)
    funcionarios = cur.fetchall()

    return render_template("funcionarios.html",funcionarios=funcionarios)

@app.route('/crud-funcionario', methods=['POST'])
def crud_funcionario():
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    data = request.get_json()

    print(data)

    matricula = data.get('matricula')
    nome = data.get('nome')
    setor = data.get('setor')
    data_admissao = data.get('dataAdmissao')
    acao = data.get('acao')  # Ex: 'create', 'update', 'delete'
    matricula_anterior = data.get('matriculaAnterior')
    ativo_funcionario = data.get('habilitarFuncionario') == 'true'
    # nome_anterior = data.get('nome_anterior')
    
    if acao == 'update':
        if matricula_anterior != matricula:
            print("Matrículas diferentes")
            query_consulta = """select * from requisicao.funcionarios where matricula = %s"""
            cur.execute(query_consulta,(matricula,))
            data = cur.fetchall()
            if len(data) > 0:
                message = 'Matrícula ja existe'
                return jsonify({'message': message}), 400
            # update nas outras tabelas que puxam a matrícula
            sql_concat = f"select concat(matricula, ' - ', nome) from requisicao.funcionarios where matricula = '{matricula_anterior}'"
            cur.execute(sql_concat)
            data = cur.fetchall()
            solicitante_anterior = data[0][0]
            solicitante = f'{matricula} - {nome.strip()}'
            atualizar_dados_pos_troca_matricula(matricula, nome, setor, data_admissao, matricula_anterior, ativo_funcionario, solicitante, solicitante_anterior)
            return jsonify({'message': 'Funcionário atualizado com sucesso'}), 200
        # if nome != nome_anterior:
        #     #trocar o padrao solicitacoes, pois ele concatena o funcionario recebe com (matricula - nome)
        #     print('Apenas nomes diferentes')
        #     sql_concat = f"select concat(matricula, ' - ', nome) from requisicao.funcionarios where matricula = '{matricula_anterior}'"
        #     cur.execute(sql_concat)
        #     data = cur.fetchall()
        #     solicitante_anterior = data[0][0]
        #     solicitante = f'{matricula} - {nome.strip()}'
        #     atualizar_dados_pos_troca_nome(matricula, nome, setor, data_admissao, matricula_anterior, ativo_funcionario, solicitante, solicitante_anterior)
        #     return jsonify({'message': 'Funcionário atualizado com sucesso'}), 200 
    
        crudFuncionario.update_funcionario(matricula, nome, setor, data_admissao, matricula_anterior, ativo_funcionario)
        return jsonify({'message': 'Funcionário atualizado com sucesso'}), 200
    elif acao == 'delete':
        crudFuncionario.desabilitar_funcionario(matricula)
        return jsonify({'message': 'Funcionário desabilitado com sucesso'}), 200
    else:
        return jsonify({'message': 'Ação inválida'}), 400
    
@app.route('/edit-item-padrao',methods=['POST'])
def edit_item_padrao():

    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                        password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        dados = request.get_json()

        equipamento_edit = dados['equipamento_edit']
        solicitante_edit = dados['solicitante_edit']
        nome_padrao = dados['nome_padrao_edit']
        equipamento_anterior_edit = dados['equipamento_anterior_edit']
        funcionario = dados['funcionario_edit']
        quantidade = dados['quantidade_edit']
        observacao = dados['observacao_edit']

        print(equipamento_edit,solicitante_edit,nome_padrao, equipamento_anterior_edit, funcionario)

        query = """ UPDATE sistema_epi.padrao_solicitacao SET codigo_item = %s, quantidade = %s, observacao = %s WHERE matricula_solicitante = %s AND nome = %s AND codigo_item = %s AND funcionario_recebe = %s """
        
        values = (equipamento_edit, quantidade, observacao, solicitante_edit, nome_padrao, equipamento_anterior_edit, funcionario)

        cur.execute(query,values)

        conn.commit()

    except Exception as e:
        conn.rollback()  # ❌ Algo deu errado, desfaz tudo
        print("Erro na transação:", e)

    finally:
        cur.close()
        conn.close()

    return jsonify("Sucesso")
    
def atualizar_dados_pos_troca_matricula(matricula, nome, setor, data_admissao, matricula_anterior, ativo_funcionario, solicitante, solicitante_anterior):
    conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
                            password=DB_PASS, host=DB_HOST)
    cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

    try:
        # Início da transação
        #Atualiza a matricula de funcionario
        cur.execute("UPDATE requisicao.funcionarios SET matricula = %s, nome = %s, setor = %s, data_admissao = %s, ativo = %s WHERE matricula = %s",
                        (matricula, nome, setor, data_admissao, ativo_funcionario, matricula_anterior))
        #Atualiza a matricula nas solicitacoes
        cur.execute("UPDATE sistema_epi.tb_solicitacoes SET funcionario_recebe = %s WHERE funcionario_recebe = %s",(matricula, matricula_anterior))
        #Atualiza a tabela de padrao_solicitacao
        cur.execute("UPDATE sistema_epi.padrao_solicitacao SET funcionario_recebe = %s WHERE funcionario_recebe = %s", (solicitante, solicitante_anterior))

        conn.commit()  # ✅ Tudo certo, confirma as mudanças

    except Exception as e:
        conn.rollback()  # ❌ Algo deu errado, desfaz tudo
        print("Erro na transação:", e)

    finally:
        cur.close()
        conn.close()

# def atualizar_dados_pos_troca_nome(matricula, nome, setor, data_admissao, matricula_anterior, ativo_funcionario, solicitante, solicitante_anterior):
#     conn = psycopg2.connect(dbname=DB_NAME, user=DB_USER,
#                             password=DB_PASS, host=DB_HOST)
#     cur = conn.cursor(cursor_factory=psycopg2.extras.DictCursor)

#     try:
#         # Início da transação
#         #Atualiza a matricula de funcionario
#         cur.execute("UPDATE requisicao.funcionarios SET matricula = %s, nome = %s, setor = %s, data_admissao = %s, ativo = %s WHERE matricula = %s",
#                         (matricula, nome, setor, data_admissao, ativo_funcionario, matricula_anterior))
#         #Atualiza a tabela de padrao_solicitacao
#         cur.execute("UPDATE sistema_epi.padrao_solicitacao SET funcionario_recebe = %s WHERE funcionario_recebe = %s", (solicitante, solicitante_anterior))

#         conn.commit()  # ✅ Tudo certo, confirma as mudanças

#     except Exception as e:
#         conn.rollback()  # ❌ Algo deu errado, desfaz tudo
#         print("Erro na transação:", e)

#     finally:
#         cur.close()
#         conn.close()



if __name__ == '__main__':
    app.run(debug=True)